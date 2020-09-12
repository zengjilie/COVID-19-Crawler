import requests
from lxml import etree
import json

import openpyxl

#1. 发送请求
url = 'https://voice.baidu.com/act/newpneumonia/newpneumonia'
response = requests.get(url)

#2. 获得数据
response = response.text
#response.text获得链接源代码


#3. 解析数据
#etree.HTML生成xml树对象.etree可以自动修复html文档,让xpath解析
html = etree.HTML(response)


#4. 筛选数据
result = html.xpath('//script[@type="application/json"]/text()')
#print(result)
result = result[0]


#4-1 进一步筛选数据
#将xml格式转成dict格式
result = json.loads(result)
# print(result)
result_in = result['component'][0]['caseList']
result_out = result['component'][0]['globalList']
summary_in = result['component'][0]['summaryDataIn']
summary_out = result['component'][0]['summaryDataOut']

print(result_out)

# print(result_out)

#5. 将数据写入xlsx文件
#建立工作簿
wb = openpyxl.Workbook()

# A 建立国内疫情工作表
#国内数据(10)
# area-->省 份/直 辖 市 / 特 别 行 政 区 等
# city --> 城 市
# confirmed --> 累 计 确 诊 人 数
# died --> 死 亡 人 数
# crued --> 治 愈 人 数
# confirmedRelative --> 累 计 确 诊 的 增 呈  
# cruedRelative --> 治 愈 的 增 蛋  
# curConfirm --> 现 有 确 诊 人 数  
# curConfirmRelative --> 现 有 确 诊 的 增 呈  
# diedRelative --> 死 亡 的 增 暈
ws = wb.active
ws.title = "国内疫情"
ws.append(['province','confirmed','died','crued','curConfirm','confirmedRelative','diedRelative','curedRelative','curConfirmRelative'])
# print(result)
for each in result_in:
    tempt_list = [each['area'],each['confirmed'],each['died'],each['crued'],each['curConfirm'],
    each['confirmedRelative'],each['diedRelative'],each['curedRelative'],each['curConfirmRelative']]
    for i in range(len(tempt_list)):
        if tempt_list[i] == '':
            tempt_list[i] = '0'
    ws.append(tempt_list)

# B 建立国外疫情工作表
# 国外疫情数据(6)
# country-->国家
# died-->死亡人数
# confirmed 确诊人数
# crued 治愈人数
# curConfirm 现有确诊人数
# confirmedRelative 现有确诊的增长
ws_out = wb.create_sheet('国外疫情')
ws_out.append(['country','confirmed','died','crued','curConfirm','confirmedRelative'])
for each in result_out[0:6]:
    print(each)
    print('************************************')
    
    for country in each['subList']:
        tempt_list = [country['country'],country['confirmed'],country['died'],country['crued'],country['curConfirm'],
        country['confirmedRelative']]
        for i in range(len(tempt_list)):
            if tempt_list[i] == '':
                tempt_list[i] = '0'
        ws_out.append(tempt_list)


# C 建立国内疫情整体疫情工作表
#国内总体数据(20)
# confirmed 累计确诊
# died 累计死亡
# cured 累计治愈
# asymptomatic 无症状 
# asymptomaticRelative 无症状增量
# unconfirmed 现有疑似
# relativeTime 数据更新时间
# confirmedRelative 累计确诊增量
# unconfirmedRelative 现有疑似增量
# curedRelative 累计治愈增量
# diedRelative 累计死亡增量
# icu 现有重症
# icuRelative 现有重症增量
# overseasInput 境外输入
# unOverseasInputCumulative 本土累计
# overseasInputRelative 境外输入增量
# unOverseasInputNewAdd 本土累计新增
# curConfirm 现有确诊
# curConfirmRelative 现有确诊增量
# icuDisable
ws_summary_in = wb.create_sheet('国内疫情整体数据')
ws_summary_in.append(['confirmed','died','cured','overseasInput','confirmedRelative','diedRelative','curedRelative','overseasInputRelative'])

tempt_list_in = [summary_in['confirmed'],summary_in['died'],summary_in['cured'],summary_in['overseasInput'],summary_in['confirmedRelative'],
summary_in['diedRelative'],summary_in['curedRelative'],summary_in['overseasInputRelative']]
print(tempt_list_in)
ws_summary_in.append(tempt_list_in)
    
# D 建立国外疫情整体疫情工作表
#国外总体数据(9)
# confirmed 累计确诊
# died 累计死亡
# curConfirm 现存确诊
# cured 累计治愈
# relativeTime 数据更新时间
# confirmedRelative 累计确诊增量
# unconfirmedRelative 现有疑似增量
# curedRelative 累计治愈增量
# diedRelative 累计死亡增量

ws_summary_out = wb.create_sheet('国外疫情整体数据')
ws_summary_out.append(['confirmed','died','curConfirm','cured','confirmedRelative','curedRelative','diedRelative','curConfirmRelative'])
tempt_list_out = [summary_out['confirmed'],summary_out['died'],summary_out['curConfirm'],summary_out['cured'],summary_out['confirmedRelative'],
summary_out['curedRelative'],summary_out['diedRelative'],summary_out['curConfirmRelative']]
# print(tempt_list_out)
ws_summary_out.append(tempt_list_out)

wb.save('./data.xlsx')







